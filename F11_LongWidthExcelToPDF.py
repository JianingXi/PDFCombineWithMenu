import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pythoncom
import time
from win32com.client import DispatchEx, constants

def detect_header_rows(df, max_rows=2):
    header_count = 0
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i].astype(str)
        if all(cell.isalpha() or cell.replace(' ', '').isalpha() for cell in row):
            header_count += 1
        else:
            break
    return header_count

def split_columns_by_width(df, max_total_width=100):
    col_widths = df.astype(str).apply(lambda col: col.str.len().max()).fillna(0).clip(5, 40)
    groups, cur, cur_w = [], [], 0
    for col, w in col_widths.items():
        if cur and cur_w + w > max_total_width:
            groups.append(cur)
            cur, cur_w = [col], w
        else:
            cur.append(col)
            cur_w += w
    if cur:
        groups.append(cur)
    return groups

def create_split_workbook(df_data, grouped_cols, output_excel_path, df_header=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    thin = Border(*(Side(style='thin') for _ in range(4)))
    align = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True)

    for idx, cols in enumerate(grouped_cols, start=1):
        ws = wb.create_sheet(title=f"Part_{idx}")
        if df_header is not None and not df_header.empty:
            for r, _ in df_header.iterrows():
                for c, col in enumerate(cols, start=1):
                    val = df_header.at[r, col]
                    cell = ws.cell(row=r+1, column=c, value=val)
                    cell.font = header_font; cell.border = thin; cell.alignment = align

        start_row = (len(df_header) if df_header is not None else 0) + 1
        for r_off, row in enumerate(dataframe_to_rows(df_data[cols], index=False, header=False)):
            r = start_row + r_off
            max_lines = 1
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin; cell.alignment = align
                txt = str(val) if val is not None else ""
                col_letter = openpyxl.utils.get_column_letter(c)
                col_w = ws.column_dimensions[col_letter].width or 10
                lines = max(1, -(-len(txt) // int(col_w * 1.2)))
                max_lines = max(max_lines, lines)
            ws.row_dimensions[r].height = max_lines * 15

        for c_idx, col in enumerate(df_data[cols].columns, start=1):
            max_len = df_data[col].astype(str).str.len().max()
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = min(max_len+5, 50)

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize   = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage    = True
        ws.page_setup.fitToWidth   = 1
        ws.page_setup.fitToHeight  = False

    wb.save(output_excel_path)
    print(f"✅ 拆分并保存为：{output_excel_path}")

def export_excel_to_pdf(excel_path, pdf_path):
    pythoncom.CoInitialize()
    excel = wb = None
    try:
        excel = DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(excel_path, ReadOnly=True)
        for sh in wb.Sheets:
            sh.PageSetup.Orientation    = constants.xlLandscape
            sh.PageSetup.PaperSize      = constants.xlPaperA4
            sh.PageSetup.Zoom           = False
            sh.PageSetup.FitToPagesWide = 1
            sh.PageSetup.FitToPagesTall = False

        wb.ExportAsFixedFormat(constants.xlTypePDF, pdf_path)
        print(f"📄 已导出 PDF：{pdf_path}")
    except Exception as e:
        print(f"❌ 导出 PDF 失败：{e}")
    finally:
        if wb:
            try: wb.Close(False)
            except: pass
        if excel:
            try: excel.Quit()
            except: pass
        pythoncom.CoUninitialize()

def long_width_excel_to_pdf(input_excel, output_pdf):
    df_full = pd.read_excel(input_excel, header=None, dtype=str)
    header_rows = detect_header_rows(df_full, max_rows=2)
    print(f"检测到表头行数：{header_rows}")

    df_header = df_full.iloc[:header_rows].reset_index(drop=True)
    df_data   = df_full.iloc[header_rows:].reset_index(drop=True)

    # 生成唯一临时文件
    base, _   = os.path.splitext(input_excel)
    temp_excel = f"{base}_split_{int(time.time())}.xlsx"

    grouped = split_columns_by_width(df_data)
    create_split_workbook(df_data, grouped, temp_excel, df_header=df_header)

    export_excel_to_pdf(temp_excel, output_pdf)

    # 删除临时文件
    try:
        os.remove(temp_excel)
        print(f"🗑️ 已删除临时文件：{temp_excel}")
    except Exception as e:
        print(f"⚠️ 无法删除临时文件：{e}")

if __name__ == "__main__":
    input_excel      = r"C:\Reformatted_A4_Printable_AdjustedRowHeight.xlsx"
    output_pdf       = r"C:\final_output_dispatch.pdf"

    long_width_excel_to_pdf(input_excel, output_pdf)
